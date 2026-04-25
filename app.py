from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, abort
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from functools import wraps
from pathlib import Path
from datetime import datetime, timedelta
import sqlite3, json, os, io

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / 'service_challan.db'
app = Flask(__name__)
app.secret_key = os.environ.get('APP_SECRET_KEY', 'change-this-secret-key')
app.config.update(
    REMEMBER_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=30)
)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

ROLE_PERMS = {
    'admin': {'challan', 'approve', 'audit', 'export', 'import'},
    'maker': {'challan', 'export', 'import'},
    'checker': {'approve', 'audit', 'export'},
    'viewer': {'export'}
}

class User(UserMixin):
    def __init__(self, id, username, password_hash, role, is_active=1):
        self.id = str(id)
        self.username = username
        self.password_hash = password_hash
        self.role = role
        self.active = bool(is_active)

    @property
    def is_active(self):
        return self.active


def conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def log(action, details=None):
    actor = current_user.username if getattr(current_user, 'is_authenticated', False) else 'system'
    with conn() as c:
        c.execute(
            'INSERT INTO audit_logs (action, username, details) VALUES (?,?,?)',
            (actor if False else action, actor, json.dumps(details, ensure_ascii=False) if details else None)
        )
        c.commit()


def role_required(*roles):
    def deco(f):
        @wraps(f)
        def inner(*a, **kw):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role not in roles:
                abort(403)
            return f(*a, **kw)
        return inner
    return deco


@login_manager.user_loader
def load_user(user_id):
    with conn() as c:
        r = c.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    return User(r['id'], r['username'], r['password_hash'], r['role'], r['is_active']) if r else None


def init_db():
    with conn() as c:
        c.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS service_challans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            challan_date TEXT,
            challan_no TEXT UNIQUE,
            purpose TEXT,
            challan_to TEXT,
            ship_to TEXT,
            destination TEXT,
            part_issue_deptt TEXT,
            machine_no TEXT,
            model TEXT,
            remarks TEXT,
            received_status TEXT,
            received_issued_date TEXT,
            received_by_to TEXT,
            conditions TEXT,
            status TEXT DEFAULT 'Pending Approval',
            maker_username TEXT,
            checker_username TEXT,
            approved_at TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS service_challan_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            challan_id INTEGER,
            item_name TEXT,
            item_code TEXT,
            qty REAL,
            rate REAL,
            total REAL,
            FOREIGN KEY(challan_id) REFERENCES service_challans(id) ON DELETE CASCADE
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT,
            username TEXT,
            details TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )''')
        defaults = [
            ('admin', 'Admin@12345', 'admin'),
            ('maker1', 'Maker@12345', 'maker'),
            ('checker1', 'Checker@12345', 'checker'),
            ('viewer1', 'Viewer@12345', 'viewer')
        ]
        for username, password, role in defaults:
            if not c.execute('SELECT 1 FROM users WHERE username=?', (username,)).fetchone():
                c.execute('INSERT INTO users (username,password_hash,role) VALUES (?,?,?)', (username, generate_password_hash(password), role))
        c.commit()


def parse_items(form):
    names = form.getlist('item_name[]')
    codes = form.getlist('item_code[]')
    qtys = form.getlist('qty[]')
    rates = form.getlist('rate[]')
    items = []
    for i in range(len(names)):
        item_name = (names[i] or '').strip()
        item_code = (codes[i] or '').strip()
        qty = float(qtys[i] or 0)
        rate = float(rates[i] or 0)
        if item_name or item_code or qty or rate:
            items.append({
                'item_name': item_name,
                'item_code': item_code,
                'qty': qty,
                'rate': rate,
                'total': round(qty * rate, 2)
            })
    return items


def challan_totals(challan_id):
    with conn() as c:
        row = c.execute('SELECT IFNULL(SUM(total),0) grand_total, IFNULL(SUM(qty),0) total_qty, COUNT(*) item_count FROM service_challan_items WHERE challan_id=?', (challan_id,)).fetchone()
    return dict(row)


def fetch_challan_rows(where='', params=()):
    query = f'''
        SELECT sc.*,
               IFNULL(SUM(sci.qty),0) as total_qty,
               IFNULL(SUM(sci.total),0) as grand_total,
               COUNT(sci.id) as item_count
        FROM service_challans sc
        LEFT JOIN service_challan_items sci ON sci.challan_id = sc.id
        {where}
        GROUP BY sc.id
        ORDER BY sc.id DESC
    '''
    with conn() as c:
        return c.execute(query, params).fetchall()


@app.context_processor
def inject_helpers():
    return {'now': datetime.now, 'ROLE_PERMS': ROLE_PERMS}


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        u = request.form.get('username', '').strip()
        p = request.form.get('password', '')
        with conn() as c:
            r = c.execute('SELECT * FROM users WHERE username=?', (u,)).fetchone()
        if not r or not check_password_hash(r['password_hash'], p):
            flash('Invalid credentials', 'error')
            return render_template('login.html')
        login_user(User(r['id'], r['username'], r['password_hash'], r['role'], r['is_active']))
        session.permanent = True
        log('LOGIN', {'role': r['role']})
        return redirect(url_for('dashboard'))
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    log('LOGOUT')
    logout_user()
    session.clear()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def dashboard():
    with conn() as c:
        stats = {
            'challans': c.execute('SELECT COUNT(*) c FROM service_challans').fetchone()['c'],
            'pending': c.execute("SELECT COUNT(*) c FROM service_challans WHERE status='Pending Approval'").fetchone()['c'],
            'approved': c.execute("SELECT COUNT(*) c FROM service_challans WHERE status='Approved'").fetchone()['c'],
            'value': c.execute('SELECT IFNULL(SUM(total),0) s FROM service_challan_items').fetchone()['s'],
        }
        recent = fetch_challan_rows('')[:8]
    return render_template('dashboard.html', stats=stats, recent=recent)


@app.route('/challans')
@login_required
def challans():
    q = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()
    destination = request.args.get('destination', '').strip()
    params = []
    clauses = []
    if q:
        clauses.append('(sc.challan_no LIKE ? OR sc.challan_to LIKE ? OR sc.ship_to LIKE ? OR sc.machine_no LIKE ? OR sc.model LIKE ?)')
        like = f'%{q}%'
        params.extend([like, like, like, like, like])
    if status:
        clauses.append('sc.status=?')
        params.append(status)
    if destination:
        clauses.append('sc.destination LIKE ?')
        params.append(f'%{destination}%')
    where = 'WHERE ' + ' AND '.join(clauses) if clauses else ''
    rows = fetch_challan_rows(where, tuple(params))
    return render_template('challans.html', rows=rows, q=q, status=status, destination=destination)


@app.route('/challans/new', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'maker')
def new_challan():
    if request.method == 'POST':
        data = {k: request.form.get(k, '').strip() for k in [
            'challan_date', 'challan_no', 'purpose', 'challan_to', 'ship_to', 'destination',
            'part_issue_deptt', 'machine_no', 'model', 'remarks', 'received_status',
            'received_issued_date', 'received_by_to', 'conditions'
        ]}
        items = parse_items(request.form)
        if not data['challan_no']:
            flash('Challan number is required.', 'error')
            return render_template('challan_form.html', form=data, items=items or [{}])
        if not items:
            flash('At least one item row is required.', 'error')
            return render_template('challan_form.html', form=data, items=[{}])
        try:
            with conn() as c:
                cur = c.execute('''INSERT INTO service_challans (
                    challan_date, challan_no, purpose, challan_to, ship_to, destination,
                    part_issue_deptt, machine_no, model, remarks, received_status,
                    received_issued_date, received_by_to, conditions, maker_username
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                    data['challan_date'], data['challan_no'], data['purpose'], data['challan_to'], data['ship_to'], data['destination'],
                    data['part_issue_deptt'], data['machine_no'], data['model'], data['remarks'], data['received_status'],
                    data['received_issued_date'], data['received_by_to'], data['conditions'], current_user.username
                ))
                challan_id = cur.lastrowid
                for item in items:
                    c.execute('INSERT INTO service_challan_items (challan_id,item_name,item_code,qty,rate,total) VALUES (?,?,?,?,?,?)',
                              (challan_id, item['item_name'], item['item_code'], item['qty'], item['rate'], item['total']))
                c.commit()
            log('CHALLAN_CREATED', {'challan_no': data['challan_no'], 'items': len(items)})
            flash('Service challan created successfully.', 'success')
            return redirect(url_for('challans'))
        except sqlite3.IntegrityError:
            flash('Challan number already exists.', 'error')
    default_form = {
        'challan_date': datetime.now().strftime('%Y-%m-%d'),
        'part_issue_deptt': 'STORE DEPTT',
        'received_status': 'ISSUED'
    }
    return render_template('challan_form.html', form=default_form, items=[{}])


@app.route('/challans/<int:challan_id>')
@login_required
def challan_view(challan_id):
    with conn() as c:
        challan = c.execute('SELECT * FROM service_challans WHERE id=?', (challan_id,)).fetchone()
        items = c.execute('SELECT * FROM service_challan_items WHERE challan_id=? ORDER BY id', (challan_id,)).fetchall()
    if not challan:
        abort(404)
    totals = challan_totals(challan_id)
    return render_template('challan_view.html', challan=challan, items=items, totals=totals)


@app.route('/challans/<int:challan_id>/approve', methods=['POST'])
@login_required
@role_required('admin', 'checker')
def approve_challan(challan_id):
    with conn() as c:
        c.execute("UPDATE service_challans SET status='Approved', checker_username=?, approved_at=datetime('now','localtime') WHERE id=?", (current_user.username, challan_id))
        c.commit()
    log('CHALLAN_APPROVED', {'challan_id': challan_id})
    flash('Challan approved.', 'success')
    return redirect(url_for('approvals'))


@app.route('/challans/<int:challan_id>/reject', methods=['POST'])
@login_required
@role_required('admin', 'checker')
def reject_challan(challan_id):
    with conn() as c:
        c.execute("UPDATE service_challans SET status='Rejected', checker_username=?, approved_at=datetime('now','localtime') WHERE id=?", (current_user.username, challan_id))
        c.commit()
    log('CHALLAN_REJECTED', {'challan_id': challan_id})
    flash('Challan rejected.', 'success')
    return redirect(url_for('approvals'))


@app.route('/approvals')
@login_required
@role_required('admin', 'checker')
def approvals():
    rows = fetch_challan_rows("WHERE sc.status='Pending Approval'")
    return render_template('approvals.html', rows=rows)


@app.route('/challans/<int:challan_id>/print')
@login_required
def print_challan(challan_id):
    with conn() as c:
        challan = c.execute('SELECT * FROM service_challans WHERE id=?', (challan_id,)).fetchone()
        items = c.execute('SELECT * FROM service_challan_items WHERE challan_id=? ORDER BY id', (challan_id,)).fetchall()
    if not challan:
        abort(404)
    totals = challan_totals(challan_id)
    return render_template('print_challan.html', challan=challan, items=items, totals=totals)


@app.route('/import-excel', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'maker')
def import_excel():
    inserted, skipped = 0, 0
    if request.method == 'POST':
        f = request.files.get('excel_file')
        if not f:
            flash('Please choose an Excel file.', 'error')
            return render_template('import_excel.html')
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows_by_challan = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rec = dict(zip(headers, row))
            challan_no = str(rec.get('CHALLAN NO') or '').strip()
            if not challan_no:
                continue
            rows_by_challan.setdefault(challan_no, []).append(rec)
        with conn() as c:
            for challan_no, group in rows_by_challan.items():
                if c.execute('SELECT 1 FROM service_challans WHERE challan_no=?', (challan_no,)).fetchone():
                    skipped += 1
                    continue
                head = group[0]
                cur = c.execute('''INSERT INTO service_challans (
                    challan_date, challan_no, purpose, challan_to, ship_to, destination,
                    part_issue_deptt, machine_no, model, remarks, received_status,
                    received_issued_date, received_by_to, conditions, maker_username, status
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                    str(head.get('CHALLAN DATE') or ''), challan_no, str(head.get('PURPOSE') or ''),
                    str(head.get('CHALLAN TO ') or ''), str(head.get('SHIP TO') or ''), str(head.get('DESTINATIONS') or ''),
                    str(head.get('PART ISSUE DEPTT.') or ''), str(head.get('MACHINE NO') or ''), str(head.get('MODEL') or ''),
                    str(head.get('REMARKS') or ''), str(head.get('Unnamed: 15') or ''),
                    str(head.get('REVEIVED /IUSSED DATE ') or ''), str(head.get('RECEIVED BY / TO ') or ''),
                    str(head.get('CONDITIONS') or ''), current_user.username, 'Approved'
                ))
                challan_id = cur.lastrowid
                for item in group:
                    qty = float(item.get('QTY') or 0)
                    rate = float(item.get('RATE') or 0)
                    total = float(item.get('TOTAL') or (qty * rate))
                    c.execute('INSERT INTO service_challan_items (challan_id,item_name,item_code,qty,rate,total) VALUES (?,?,?,?,?,?)', (
                        challan_id,
                        str(item.get('ITEM NAME ') or ''),
                        str(item.get('ITEM CODE') or ''),
                        qty, rate, total
                    ))
                inserted += 1
            c.commit()
        log('EXCEL_IMPORTED', {'inserted': inserted, 'skipped': skipped})
        flash(f'Import completed. Inserted: {inserted}, Skipped existing challans: {skipped}', 'success')
    return render_template('import_excel.html')


@app.route('/export-register')
@login_required
def export_register():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Service Challan Register'
    ws.append(['CHALLAN DATE', 'CHALLAN NO', 'PURPOSE', 'CHALLAN TO', 'SHIP TO', 'DESTINATIONS', 'PART ISSUE DEPTT.', 'MACHINE NO', 'MODEL', 'ITEM NAME', 'ITEM CODE', 'QTY', 'RATE', 'TOTAL', 'REMARKS', 'RECEIVED/ISSUED', 'RECEIVED / ISSUED DATE', 'RECEIVED BY / TO', 'CONDITIONS', 'STATUS', 'MAKER', 'CHECKER'])
    with conn() as c:
        rows = c.execute('''
            SELECT sc.*, sci.item_name, sci.item_code, sci.qty, sci.rate, sci.total
            FROM service_challans sc
            LEFT JOIN service_challan_items sci ON sci.challan_id = sc.id
            ORDER BY sc.id DESC, sci.id ASC
        ''').fetchall()
    for r in rows:
        ws.append([
            r['challan_date'], r['challan_no'], r['purpose'], r['challan_to'], r['ship_to'], r['destination'], r['part_issue_deptt'],
            r['machine_no'], r['model'], r['item_name'], r['item_code'], r['qty'], r['rate'], r['total'], r['remarks'],
            r['received_status'], r['received_issued_date'], r['received_by_to'], r['conditions'], r['status'], r['maker_username'], r['checker_username']
        ])
    out = BASE_DIR / 'service_challan_register.xlsx'
    wb.save(out)
    log('REGISTER_EXPORTED')
    return send_file(out, as_attachment=True, download_name='service_challan_register.xlsx')


@app.route('/audit')
@login_required
@role_required('admin', 'checker')
def audit():
    with conn() as c:
        rows = c.execute('SELECT * FROM audit_logs ORDER BY id DESC LIMIT 300').fetchall()
    return render_template('audit.html', rows=rows)


@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', code=403, message='You do not have permission to access this page.'), 403


@app.errorhandler(404)
def missing(e):
    return render_template('error.html', code=404, message='Requested record or page was not found.'), 404


if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=False)
